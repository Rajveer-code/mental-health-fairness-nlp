"""
compile_tables.py
─────────────────
Compiles all manuscript tables (2–10) from result CSVs into a single
Excel workbook, with multi-seed mean ± std formatting for Tables 2–4.

This script is the final aggregation step before manuscript rebuilding.
It reads outputs from all upstream analysis scripts and produces one
xlsx file with one sheet per table, ready for copy-paste into the
submission-ready manuscript.

Inputs
------
outputs/results/multiseed/multiseed_results.csv
    From train_multiseed.py: per-seed metrics + summary rows.

manuscript_inputs/fairness/pairwise_auc_comparisons.csv
    From fairness_audit.py: between-model Bonferroni-corrected Z-tests (Table 5).

manuscript_inputs/fairness/di_eod_table.csv
    From code_A1_di_eod_analysis.py: DI/EOD per class (Table 6).

manuscript_inputs/fairness/temperature_scaling_results.csv
    From code_A5_temperature_scaling.py: Table 7.

outputs/results/ig_table8_update.csv
    From integrated_gradients.py: Table 8 with IG J column.

manuscript_inputs/fairness/sensitivity_drops_all_mappings.csv
    From sensitivity_analysis.py + label_sensitivity_mappings_DE.py: Table 9.

outputs/results/calibration_comparison.csv
    From calibration_comparison.py: New Table 10.

Outputs
-------
outputs/results/all_tables.xlsx
    One sheet per table. Sheet names: Table2 … Table10.

outputs/results/tables_summary.txt
    Plain-text summary of which tables were successfully generated.

Usage
-----
Run from the repository root:
    python src/compile_tables.py
"""

import os
import sys
import warnings

import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=FutureWarning)

sys.path.insert(0, os.path.dirname(__file__))
from utils import (
    MODELS,
    PLATFORMS,
    CLASSES,
    MODEL_DISPLAY,
    load_config,
)

# ── Config ─────────────────────────────────────────────────────────────────────

cfg = load_config()

RESULTS_DIR   = cfg["paths"]["results"]
MULTISEED_DIR = os.path.join(RESULTS_DIR, "multiseed")
FAIRNESS_DIR  = os.path.join("manuscript_inputs", "fairness")
OUT_XLSX      = os.path.join(RESULTS_DIR, "all_tables.xlsx")

# Column display format for mean ± std
FMT4  = "{:.4f}"     # 4 decimal places for AUC
FMT3  = "{:.3f}"     # 3 decimal places for ECE / F1
FMT4S = "{:.4f} ± {:.4f}"
FMT3S = "{:.3f} ± {:.3f}"


# ── Formatting helpers ─────────────────────────────────────────────────────────

def _fmt(mean: float, std: float | None, dp: int = 4) -> str:
    """Format a mean ± std value for the manuscript tables."""
    fmt = f"{{:.{dp}f}}"
    if std is None or np.isnan(std) or std == 0:
        return fmt.format(mean)
    return f"{fmt.format(mean)} ± {fmt.format(std)}"


def _style_header(ws, row: int, n_cols: int) -> None:
    """Apply bold header styling to a worksheet row."""
    fill   = PatternFill("solid", fgColor="1F4E79")
    font   = Font(bold=True, color="FFFFFF")
    align  = Alignment(horizontal="center", wrap_text=True)
    for col in range(1, n_cols + 1):
        cell          = ws.cell(row=row, column=col)
        cell.fill     = fill
        cell.font     = font
        cell.alignment = align


def _style_data_row(ws, row: int, n_cols: int, even: bool) -> None:
    """Apply alternating row colours for readability."""
    fgColor = "EFF3FB" if even else "FFFFFF"
    fill    = PatternFill("solid", fgColor=fgColor)
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).alignment = Alignment(
            horizontal="center", vertical="center"
        )


def _autofit_columns(ws, min_width: int = 10, max_width: int = 40) -> None:
    """Set column widths based on content."""
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=min_width,
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, min_width), max_width
        )


def _write_df_to_sheet(
    ws,
    df: pd.DataFrame,
    title: str,
    description: str = "",
) -> None:
    """Write a DataFrame to an openpyxl worksheet with consistent styling."""
    # Title row
    ws.cell(row=1, column=1).value = title
    ws.cell(row=1, column=1).font  = Font(bold=True, size=13)
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 1)
    )

    if description:
        ws.cell(row=2, column=1).value = description
        ws.cell(row=2, column=1).font  = Font(italic=True, size=9)
        ws.merge_cells(
            start_row=2, start_column=1, end_row=2,
            end_column=max(len(df.columns), 1),
        )
        header_row = 3
    else:
        header_row = 2

    # Header
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=col_idx).value = col_name
    _style_header(ws, header_row, len(df.columns))

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
        _style_data_row(ws, row_idx, len(df.columns), even=(row_idx % 2 == 0))

    _autofit_columns(ws)


# ── Table builders ─────────────────────────────────────────────────────────────

def _load_multiseed_summary() -> pd.DataFrame | None:
    """Load the summary rows (mean ± std) from multiseed_results.csv."""
    path = os.path.join(MULTISEED_DIR, "multiseed_results.csv")
    if not os.path.exists(path):
        print(f"  MISSING: {path}")
        return None
    df = pd.read_csv(path)
    if "row_type" in df.columns:
        return df[df["row_type"] == "summary"].copy()
    return df


def build_table2(ms: pd.DataFrame) -> pd.DataFrame:
    """
    Table 2 — Within-platform performance (Kaggle test set).
    Columns: Model | Accuracy | F1-macro | F1-weighted | AUC | ECE
    """
    sub = ms[ms["platform"] == "kaggle"].copy()
    rows = []
    for _, r in sub.iterrows():
        rows.append({
            "Model":       MODEL_DISPLAY.get(r["model"], r["model"]),
            "Accuracy":    _fmt(r["accuracy"],    r.get("accuracy_std"),    4),
            "F1-macro":    _fmt(r["f1_macro"],     r.get("f1_macro_std"),    3),
            "F1-weighted": _fmt(r["f1_weighted"],  r.get("f1_weighted_std"), 3),
            "AUC":         _fmt(r["auc_macro"],    r.get("auc_macro_std"),   4),
            "ECE":         _fmt(r["ece"],          r.get("ece_std"),          3),
        })
    return pd.DataFrame(rows)


def build_table3(ms: pd.DataFrame) -> pd.DataFrame:
    """
    Table 3 — Cross-platform AUC and ECE degradation.
    Rows: model × platform. Columns: AUC, ECE, ΔF1 vs Kaggle.
    """
    rows = []
    for model_key in MODELS:
        sub_k = ms[(ms["model"] == model_key) & (ms["platform"] == "kaggle")]
        auc_k = sub_k["auc_macro"].values[0] if not sub_k.empty else float("nan")

        for platform in PLATFORMS:
            sub = ms[(ms["model"] == model_key) & (ms["platform"] == platform)]
            if sub.empty:
                continue
            r   = sub.iloc[0]
            auc = r["auc_macro"]
            rows.append({
                "Model":         MODEL_DISPLAY.get(model_key, model_key),
                "Platform":      platform.capitalize(),
                "AUC":           _fmt(auc, r.get("auc_macro_std"), 4),
                "ECE":           _fmt(r["ece"], r.get("ece_std"), 3),
                "F1-macro":      _fmt(r["f1_macro"], r.get("f1_macro_std"), 3),
                "ΔAUC vs Kaggle": (
                    f"{auc - auc_k:+.4f}" if platform != "kaggle" else "—"
                ),
            })
    return pd.DataFrame(rows)


def build_table4(ms: pd.DataFrame) -> pd.DataFrame:
    """
    Table 4 — Per-class F1 across platforms.
    """
    rows = []
    for _, r in ms.iterrows():
        row = {
            "Model":    MODEL_DISPLAY.get(r["model"], r["model"]),
            "Platform": r["platform"].capitalize(),
        }
        for cls in CLASSES:
            col = f"f1_{cls}"
            col_std = f"f1_{cls}_std"
            row[cls.capitalize()] = _fmt(
                r.get(col, float("nan")),
                r.get(col_std),
                3,
            )
        rows.append(row)
    return pd.DataFrame(rows)


def build_table5() -> pd.DataFrame | None:
    """Table 5 — Pairwise AUC comparisons (Bonferroni-corrected Z-tests)."""
    path = os.path.join(FAIRNESS_DIR, "pairwise_auc_comparisons.csv")
    if not os.path.exists(path):
        print(f"  MISSING: {path}")
        return None
    df = pd.read_csv(path)
    # Rename columns for clarity if needed
    rename = {"model_a": "Model A", "model_b": "Model B"}
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    return df


def build_table6() -> pd.DataFrame | None:
    """Table 6 — Disparate Impact and Equalized Odds Difference."""
    path = os.path.join(FAIRNESS_DIR, "di_eod_table.csv")
    if not os.path.exists(path):
        print(f"  MISSING: {path}")
        return None
    return pd.read_csv(path)


def build_table7() -> pd.DataFrame | None:
    """Table 7 — Temperature scaling results."""
    path = os.path.join(FAIRNESS_DIR, "temperature_scaling_results.csv")
    if not os.path.exists(path):
        # Also try results dir (code_A5 saves there)
        path2 = os.path.join(RESULTS_DIR, "fairness", "temperature_scaling_results.csv")
        if os.path.exists(path2):
            path = path2
        else:
            print(f"  MISSING: {path}")
            return None
    df = pd.read_csv(path)
    if "model" in df.columns:
        df["model"] = df["model"].map(MODEL_DISPLAY).fillna(df["model"])
    return df


def build_table8() -> pd.DataFrame | None:
    """
    Table 8 — Feature stability: Gradient Saliency J and IG J at K=10.
    Reads the ig_table8_update.csv produced by integrated_gradients.py.
    Falls back to the original Jaccard CSV if IG results are not yet available.
    """
    ig_path = os.path.join(RESULTS_DIR, "ig_table8_update.csv")
    if os.path.exists(ig_path):
        df = pd.read_csv(ig_path)
        display_cols = {
            "model_display": "Model",
            "class_name":    "Class",
            "pair":          "Platform Pair",
            "k":             "K",
            "jaccard_gradsal_k10": "Grad-Sal J (K=10)",
            "jaccard_ig_k10":      "IG J (K=10)",
            "agreement":           "Agreement",
        }
        df = df.rename(columns={k: v for k, v in display_cols.items() if k in df.columns})
        return df

    # Fallback: original Jaccard CSV without IG
    orig_path = os.path.join(FAIRNESS_DIR, "jaccard_full_analysis.csv")
    if not os.path.exists(orig_path):
        print(f"  MISSING: {ig_path} and {orig_path}")
        return None
    df = pd.read_csv(orig_path)
    df["IG J (K=10)"] = "Not yet computed"
    return df


def build_table9() -> pd.DataFrame | None:
    """Table 9 — Label-mapping sensitivity analysis."""
    path = os.path.join(FAIRNESS_DIR, "sensitivity_drops_all_mappings.csv")
    if not os.path.exists(path):
        print(f"  MISSING: {path}")
        return None
    df = pd.read_csv(path)
    if "model" in df.columns:
        df["model"] = df["model"].map(MODEL_DISPLAY).fillna(df["model"])
    return df


def build_table10() -> pd.DataFrame | None:
    """Table 10 (NEW) — Calibration comparison: baseline vs TS vs fine-tuning."""
    path = os.path.join(RESULTS_DIR, "calibration_comparison.csv")
    if not os.path.exists(path):
        print(f"  MISSING: {path}")
        return None
    df = pd.read_csv(path)

    # Build clean display version
    display_rows = []
    for _, r in df.iterrows():
        display_rows.append({
            "Model":           r.get("model_display", r.get("model", "")),
            "Platform":        str(r.get("platform", "")).capitalize(),
            "n (cal/eval)":    f"{r.get('n_cal', '?')}/{r.get('n_eval', '?')}",
            "Baseline AUC":    f"{r.get('baseline_auc', float('nan')):.4f}",
            "Baseline ECE":    f"{r.get('baseline_ece', float('nan')):.4f}",
            "TempScale ECE":   f"{r.get('tempscale_ece', float('nan')):.4f}",
            "TempScale ΔAUC":  f"{r.get('tempscale_auc_delta', 0.0):+.4f}",
            "FineTuned AUC":   f"{r.get('finetuned_auc', float('nan')):.4f}",
            "FineTuned ECE":   f"{r.get('finetuned_ece', float('nan')):.4f}",
            "FT AUC Gain":     f"{r.get('ft_auc_gain_vs_baseline', 0.0):+.4f}",
            "Verdict":         r.get("ft_vs_ts_verdict", ""),
        })
    return pd.DataFrame(display_rows)


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    print("Compiling all manuscript tables (2–10) …")
    print("=" * 60)

    ms = _load_multiseed_summary()

    table_specs: list[tuple[str, str, str, pd.DataFrame | None]] = []

    if ms is not None:
        table_specs += [
            (
                "Table2",
                "Table 2 — Within-Platform Performance (Kaggle test set)",
                "All values: mean ± std across 5 training seeds (42, 0, 1, 7, 123).",
                build_table2(ms),
            ),
            (
                "Table3",
                "Table 3 — Cross-Platform AUC and ECE",
                "Rows with ΔAUC show degradation relative to within-platform (Kaggle).",
                build_table3(ms),
            ),
            (
                "Table4",
                "Table 4 — Per-Class F1 Across Platforms",
                "Mean ± std across 5 seeds. Minority classes: Anxiety, Stress.",
                build_table4(ms),
            ),
        ]
    else:
        print(
            "  WARNING: multiseed_results.csv not found. "
            "Tables 2–4 will be skipped.\n"
            "  Run python src/train_multiseed.py first."
        )

    table_specs += [
        (
            "Table5",
            "Table 5 — Pairwise AUC Comparisons (Bonferroni-corrected Bootstrap Z-tests)",
            "Significant differences (p < 0.05 after correction) are marked.",
            build_table5(),
        ),
        (
            "Table6",
            "Table 6 — Fairness Metrics: Disparate Impact and Equalized Odds Difference",
            "DI < 0.80 = four-fifths rule violation. DI < 0.50 = severe disparity.",
            build_table6(),
        ),
        (
            "Table7",
            "Table 7 — Temperature Scaling Recalibration Results",
            "T* fitted on 10% stratified calibration split. AUC unchanged by monotone rescaling.",
            build_table7(),
        ),
        (
            "Table8",
            "Table 8 — Feature Attribution Stability (Jaccard Similarity, K=10)",
            "Both gradient saliency and Integrated Gradients reported. "
            "Random baseline J ≈ 0.0001.",
            build_table8(),
        ),
        (
            "Table9",
            "Table 9 — Label-Mapping Sensitivity Analysis",
            "AUC degradation across five label-mapping schemas (A–E).",
            build_table9(),
        ),
        (
            "Table10",
            "Table 10 (NEW) — Calibration Comparison: Temperature Scaling vs Fine-Tuning",
            "Both conditions trained on identical 10% stratified calibration split "
            "(n≈288 Twitter, n≈626 Reddit). "
            "Temperature scaling recovers calibration; fine-tuning may also recover AUC.",
            build_table10(),
        ),
    ]

    # Write to Excel
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        success_count = 0
        for sheet_name, title, description, df in table_specs:
            if df is None or df.empty:
                print(f"  SKIP {sheet_name}: no data available")
                continue

            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            ws = writer.sheets[sheet_name]

            # Write title above data
            ws.cell(row=1, column=1).value = title
            ws.cell(row=1, column=1).font  = Font(bold=True, size=12)
            ws.cell(row=2, column=1).value = description
            ws.cell(row=2, column=1).font  = Font(italic=True, size=9)

            _style_header(ws, row=3, n_cols=len(df.columns))
            for row_idx in range(4, 4 + len(df)):
                _style_data_row(ws, row_idx, len(df.columns), even=(row_idx % 2 == 0))
            _autofit_columns(ws)

            print(f"  ✓ {sheet_name}: {len(df)} rows")
            success_count += 1

    print(f"\n{'='*60}")
    print(f"Saved: {OUT_XLSX}")
    print(f"Tables written: {success_count}/{len(table_specs)}")

    # Summary file
    summary_lines = [
        "Table Compilation Summary",
        "=" * 40,
    ]
    for sheet_name, title, _, df in table_specs:
        status = f"{len(df)} rows" if df is not None and not df.empty else "MISSING"
        summary_lines.append(f"{sheet_name:<10} {status:<12} {title[:50]}")

    summary_path = os.path.join(RESULTS_DIR, "tables_summary.txt")
    with open(summary_path, "w") as fh:
        fh.write("\n".join(summary_lines))
    print(f"Summary: {summary_path}")

    print(f"\nNext step: python src/generate_figures_v2.py")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
